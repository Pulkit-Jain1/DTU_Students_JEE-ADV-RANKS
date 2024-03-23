from PyPDF2 import PdfReader
from tabulate import tabulate
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from openpyxl import Workbook


reader_beta = PdfReader("2023_1.pdf")
reader_alpha = PdfReader("2023_.pdf")
nbeta = len(reader_beta.pages)
nalpha = len(reader_alpha.pages)
print("alpha is adv pdf and no. of pages in it are", nalpha)
print("beta is dtu pdf and no. of pages in it are",nbeta)



# only application number 
final_list = []
for i in range(0, nbeta):
    page = reader_beta.pages[i]
    rawdata = page.extract_text()
    my_list = rawdata.split()
    final_list += [word for word in my_list if len(word) == 12 and word.startswith('23')]

print(len(final_list))



#only advance app no of 3.5 lakh
final_list_alpha = []
for i in range(548,nalpha-1):
    page1 = reader_alpha.pages[i]
    rawdata1 = page1.extract_text()
    my_list1 = rawdata1.split()
    # final_list_alpha += [word for word in my_list1 if len(word) == 12 and word.startswith('23')]
    final_list_alpha += [word for word in my_list1 if (len(word) == 12 and word.startswith('23')) or (len(word) == 9 and word.startswith('23'))]

print(final_list_alpha)




# list having app no and then adv roll number
new_list = []
for roll_number_1 in final_list:
    if roll_number_1 in final_list_alpha:
        roll_number_2_index = final_list_alpha.index(roll_number_1) - 1
        new_list.append(roll_number_1)
        new_list.append(final_list_alpha[roll_number_2_index])
    else:
        new_list.append(roll_number_1)
        new_list.append(0)

print(new_list)
print(len(new_list))
ll = new_list #FINAL LIST TILL NOW HAVING MAINS ROLL NUMBER AND AFTER THAT CORRESPONDING ADVCANCE RANK



#crl
crl_list = []
for i in range(50,131):
    page = reader_alpha.pages[i]
    rawdata = page.extract_text()
    my_list = rawdata.split()
    crl_list += [s for i, s in enumerate(my_list) if s.isdigit() and (i == 0 or my_list[i-1] != "page")]
    
# removing last elements as there is a wierd result becuase of it 
crl_list = crl_list[:-1]

print(crl_list[0])
print(crl_list[1])
print(crl_list[-6])
print(crl_list[-5])
print(crl_list[-4])
print(crl_list[-3])
print(crl_list[-2])
print(crl_list[-1])
print(len(crl_list))



#EWS
ews_list = []
for i in range(131,148):
    page = reader_alpha.pages[i]
    rawdata = page.extract_text()
    my_list = rawdata.split()
    ews_list += [s for i, s in enumerate(my_list) if s.isdigit() and (i == 0 or my_list[i-1] != "page")]

ews_list = ews_list[:-1]
print(ews_list[0])
print(ews_list[1])
print(ews_list[-4])
print(ews_list[-3])
print(ews_list[-2])
print(ews_list[-1])
print(len(ews_list))



#OBC
obc_list = []
for i in range(148,176):
    page = reader_alpha.pages[i]
    rawdata = page.extract_text()
    my_list = rawdata.split()
    obc_list += [s for i, s in enumerate(my_list) if s.isdigit() and (i == 0 or my_list[i-1] != "page")]

obc_list = obc_list[:-1]
print(obc_list[0])
print(obc_list[1])
print(obc_list[-4])
print(obc_list[-3])
print(obc_list[-2])
print(obc_list[-1])
print(len(obc_list))



#SC
sc_list = []
for i in range(176,192):
    page = reader_alpha.pages[i]
    rawdata = page.extract_text()
    my_list = rawdata.split()
    sc_list += [s for i, s in enumerate(my_list) if s.isdigit() and (i == 0 or my_list[i-1] != "page")]

sc_list = sc_list[:-1]
print(sc_list[0])
print(sc_list[1])
print(sc_list[-4])
print(sc_list[-3])
print(sc_list[-2])
print(sc_list[-1])
print(len(sc_list))



#ST
st_list = []
for i in range(192,198):
    page = reader_alpha.pages[i]
    rawdata = page.extract_text()
    my_list = rawdata.split()
    st_list += [s for i, s in enumerate(my_list) if s.isdigit() and (i == 0 or my_list[i-1] != "page")]

st_list = st_list[:-1]
print(st_list[0])
print(st_list[1])
print(st_list[-4])
print(st_list[-3])
print(st_list[-2])
print(st_list[-1])
print(len(st_list))



final_list = []
for i in range(0, len(new_list), 2):
    old_roll_number = new_list[i]
    new_roll_number = new_list[i+1]
    
    # Retrieve rank in crl exam
    crl_index = crl_list.index(new_roll_number) if new_roll_number in crl_list else -1
    crl_rank = crl_list[crl_index + 1] if crl_index != -1 else 0
    
    # Retrieve rank in ews exam
    ews_index = ews_list.index(new_roll_number) if new_roll_number in ews_list else -1
    ews_rank = ews_list[ews_index + 1] if ews_index != -1 else 0
    
    # Retrieve rank in obc exam
    obc_index = obc_list.index(new_roll_number) if new_roll_number in obc_list else -1
    obc_rank = obc_list[obc_index + 1] if obc_index != -1 else 0
    
    # Retrieve rank in sc exam
    sc_index = sc_list.index(new_roll_number) if new_roll_number in sc_list else -1
    sc_rank = sc_list[sc_index + 1] if sc_index != -1 else 0
    
    # Retrieve rank in st exam
    st_index = st_list.index(new_roll_number) if new_roll_number in st_list else -1
    st_rank = st_list[st_index + 1] if st_index != -1 else 0
    
    # Append data for the current student to the final list
    final_list.append([old_roll_number, new_roll_number, crl_rank, ews_rank, obc_rank, sc_rank, st_rank])


print(len(final_list))
print(final_list)




# this created a list with app number name seprated by elements and roll no

# def process_data(data):
#     result = []
#     i = 0
#     while i < len(data):
#         if data[i].startswith("23") and len(data[i]) == 12:
#             result.append(data[i])
#             i += 1
#             while data[i] not in {"Bio-Technology", "Chemical", "Sec-1","Sec-2","Sec-3","Sec-4","Sec-5","Sec-6","Sec-7","Sec-8","Environmental", "Mechanical", "Production"}:
#                 result.append(data[i])
#                 i += 1
#         else:
#             i += 1
#     return result

name = []
for i in range(0, nbeta):
    page = reader_beta.pages[i]
    rawdata = page.extract_text()
    my_list = rawdata.split()
    name +=   my_list#process_data(my_list)


# print(name)
# print(name[498])
# print(name[499])
# print(name[500])
# print(name[501])
# print(name[502])
# print(name[-5])
# print(name[-4])
# print(name[-3])
# print(name[-2])
# print(name[-1])
# print(len(name))





# this created a list with app number combined name and roll number

# def process_list(data):

#   new_list = []
#   intermediate = []
#   for item in data:
#     if item.startswith("23") and len(item) == 12:
#       # Found a 12-letter element starting with "23"
#       if intermediate:
#         # Combine intermediate elements with spaces
#         new_list.append(" ".join(intermediate))
#         intermediate = []
#       new_list.append(item)
#     elif item.startswith("23/"):
#       # Found an element starting with "23/"
#       if intermediate:
#         # Combine intermediate elements with spaces before this element
#         new_list.append(" ".join(intermediate))
#         intermediate = []
#       new_list.append(item)
#     else:
#       # Other elements, add to intermediate list
#       intermediate.append(item)
  
#   # Add any remaining intermediate elements at the end
#   if intermediate:
#     new_list.append(" ".join(intermediate))
  
#   return new_list

print("process")
def process_list(data):
    new_list = []
    flag = False
    repeat = False
    n = len(data)
    i = 0
    while data[i] != "23/SE/181":
        # print(len(new_list),'\n')
        if data[i].startswith("23") and len(data[i]) == 12:
            new_list.append(data[i])
            flag = True
            repeat = False
            i += 1
        elif data[i].startswith("23/") and flag == True:
            new_list.append(data[i])
            flag = False
            i+=1
        elif flag == True:
            if repeat == False:
                new_list.append(data[i])
                repeat = True
                i +=1 
            elif repeat == True:
                new_list[-1] = new_list[-1] + " " + data[i]
                del data[i]
        else:
            i += 1
            continue
        
    return new_list


nam = process_list(name)
nam.append("23/SE/181")
print(len(nam))
print(nam[5865])
print(nam[5866])
print(nam[5867])
print(nam[5868])
print(nam[5869])
print(nam[5870])
print(nam[5871])
print(nam[5872])
print(nam[5873])
print(nam[5874])
print(nam[5875])
print(nam[5876])
print(nam[5877])
print(2493*3)


# wb = Workbook()
# ws = wb.active

# # Adding headers for the first set of data
# headers = ["a", "b", "c"]
# ws.append(headers)

# # Adding student data for the first set
# for i in nam:
#     ws.append(i)

# wb.save("test3.xlsx")

#final code to uncomment 
ff =[]
length_final_list = len(final_list)
i = 0
j = 0
while(i < length_final_list):
   final_list[i].insert(0,nam[j+2])
   final_list[i].insert(0,nam[j+1])
   i += 1
   j=j+3
#    ff.append(nam[j+1])
#    ff.append(nam[j+2])
#    ff.append(final_list[i])
#    ff.append(final_list[i+1])
#    ff.append(final_list[i+2])
#    ff.append(final_list[i+3])
#    ff.append(final_list[i+4])
#    ff.append(final_list[i+5])
#    ff.append(final_list[i+6])
#    i = i + 7
#    j = j + 3 
# print(final_list)
print(len(final_list))
print(final_list[-5])
print(final_list[-4])
print(final_list[-3])
print(final_list[-2])
print(final_list[-1])


# final code to uncomment 
wb = Workbook()
ws = wb.active

# Adding headers for the first set of data
headers = ["name", "roll", "mains", "adv", "crl", "ews", "obc", "sc", "st"]
ws.append(headers)

# Adding student data for the first set
for i in final_list:
    ws.append(i)

wb.save("test3.xlsx")





# # Adding headers for the second set of data
# headers2 = ["mains2", "name", "roll"]
# ws.append(headers2)

# # Adding names as rows for the second set
# for j in nam:
#     ws.append(j)
    

# wb = Workbook()
# ws = wb.active

# # Adding headers
# headers = ["mains", "adv", "crl", "ews", "obc", "sc", "st", "mains", "name", "roll"]
# ws.append(headers)

# # Adding student data
# for i in range(len(final_list)):
#     student_row = final_list[i]
#     additional_row = nam[i] if i < len(nam) else [-1, -1, -1]  # If additional data runs out, fill with zeros
#     ws.append(student_row + additional_row)

# # Saving the workbook
# wb.save("combined_data.xlsx")








# def combine_student_data(final_list, nam):
    




# ff = []
# ff = combine_student_data(final_list, nam)

# print(ff)
# print(len(ff))
# print(ff[-8])
# print(ff[-7])
# print(ff[-6])
# print(ff[-5])
# print(ff[-4])
# print(ff[-3])
# print(ff[-2])
# print(ff[-1])




# #creating table
# headers = ["Old Roll", "New Roll", "CRL Rank", "EWS Rank", "OBC Rank", "SC Rank", "ST Rank"]
# print(tabulate(final_list, headers=headers))
# pdf_filename = "student_ranks1.pdf"
# doc = SimpleDocTemplate(pdf_filename, pagesize=letter)
# table_data = []
# for row in final_list:
#     table_data.append(row)
# table_style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
#                           ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
#                           ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#                           ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#                           ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
#                           ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
#                           ('GRID', (0, 0), (-1, -1), 1, colors.black)])
# table = Table(table_data)
# table.setStyle(table_style)
# doc.build([table])








